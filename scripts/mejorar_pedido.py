"""
Mejora pedido-mayorista.xlsx:
- Encabezado azul, blanco, en negrita
- Resalta en rojo las filas con faltante > 0
- Fórmulas SUM en fila de totales
- Códigos como texto, anchos de columna ajustados

Uso: python mejorar_pedido.py pedido-mayorista-24-5-2026.xlsx
"""

import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, CellIsRule


def mejorar(path: str) -> None:
    wb = load_workbook(path)
    ws = wb["Pedido"]

    # Detectar última fila de datos (buscar primera fila vacía después de row 2)
    last_data_row = 1
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[1].value:  # columna B (Descripción) tiene valor
            last_data_row = row[0].row
        else:
            break

    # ---- Encabezado ----
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 30

    # ---- Datos: bordes y alineación ----
    for row in ws.iter_rows(min_row=2, max_row=last_data_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            if cell.column_letter in ("A", "C", "D", "E"):
                cell.alignment = center

    # Código como texto
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=1).number_format = "@"

    # ---- Totales ----
    total_row = last_data_row + 2
    ws.cell(row=total_row, column=1, value=f"TOTAL — {last_data_row - 1} items")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{last_data_row})")
    ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{last_data_row})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{last_data_row})")

    total_fill = PatternFill("solid", fgColor="F2F2F2")
    bold = Font(bold=True)
    for cell in ws[total_row]:
        cell.fill = total_fill
        cell.font = bold
        cell.border = border

    # ---- Formato condicional ----
    # Fila completa en rosa si faltante > 0
    ws.conditional_formatting.add(
        f"A2:E{last_data_row}",
        FormulaRule(
            formula=["$E2>0"],
            fill=PatternFill("solid", fgColor="F8CBAD"),
            font=Font(color="9C0006", bold=True),
        ),
    )
    # Celda de faltante en rojo fuerte si > 0
    ws.conditional_formatting.add(
        f"E2:E{last_data_row}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor="C00000"),
            font=Font(color="FFFFFF", bold=True),
        ),
    )

    # ---- Anchos de columna y panel ----
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A2"

    out = path.replace(".xlsx", "-mejorado.xlsx")
    wb.save(out)
    print(f"Guardado: {out}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python mejorar_pedido.py pedido-mayorista.xlsx")
        sys.exit(1)
    mejorar(sys.argv[1])
